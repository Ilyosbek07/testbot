from data.config import (POSTGRES_USER, POSTGRES_PASSWORD, POSTGRES_HOST, POSTGRES_PORT, POSTGRES_DB)
import openpyxl


def make_connection_string(async_fallback: bool = False) -> str:
    result = (
        f"postgresql+asyncpg://{POSTGRES_USER}:{POSTGRES_PASSWORD}@{POSTGRES_HOST}:{POSTGRES_PORT}/{POSTGRES_DB}"
    )
    if async_fallback:
        result += "?async_fallback=True"
    return result


async def get_excel_data(file_path: str) -> list:
    """
    Получение данных из excel файла
    :param file_path: путь к файлу
    :return: список словарей
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    result = []
    for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        result.append({
            row[0].value: row[1].value
        })
    return result